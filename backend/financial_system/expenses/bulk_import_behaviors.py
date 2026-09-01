import io
import re
from datetime import date as date_cls, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.http import HttpResponse
from rest_framework import status
from rest_framework.response import Response

from catalog.models import ExpenseCategory
from expenses import models
from expenses.behaviors import CreateExpenseBehavior


class BulkImportExpenseBehavior:
    """Executa criacao em lote, importacao e exclusao de parcelas de despesas."""

    MAX_ROWS = 1000

    def __init__(self, tenant_id):
        self.tenant_id = tenant_id

    def _create_from_item(self, item: dict):
        """Cria despesa(s) a partir de um item validado, respeitando quantidade/parcelamento."""
        payload = dict(item)
        payload['tenant_id'] = self.tenant_id
        behavior = CreateExpenseBehavior(data=payload)
        if behavior.is_installment and behavior.installments > 1:
            return behavior._create_installments()
        if behavior.quantity and behavior.quantity > 1:
            return behavior._create_multiple()
        return [behavior._create_single()]

    def bulk_create(self, items) -> Response:
        """Cria multiplos gastos de forma atomica apos validar suas categorias."""
        requested_ids = {item['category_id'] for item in items}
        allowed_ids = set(ExpenseCategory.objects.filter(tenant_id__in=['system', self.tenant_id], id__in=requested_ids).values_list('id', flat=True))
        invalid_ids = sorted(requested_ids - allowed_ids)
        if invalid_ids:
            return Response({'success': False, 'message': f'Categoria(s) inv?lida(s): {invalid_ids}', 'invalid_category_ids': invalid_ids}, status=status.HTTP_400_BAD_REQUEST)
        created = []
        try:
            with transaction.atomic():
                for item in items:
                    created.extend(self._create_from_item(item))
        except Exception as e:
            return Response({'success': False, 'message': f'Erro ao criar gastos em lote: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': True, 'created': len(created), 'message': f'{len(created)} gasto(s) criado(s) com sucesso'}, status=status.HTTP_201_CREATED)

    def import_template(self):
        """Gera o arquivo .xlsx modelo para importacao de gastos."""
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
        from cards.models import CreditCard
        categories = list(ExpenseCategory.objects.filter(tenant_id__in=['system', self.tenant_id]).order_by('name').values_list('name', flat=True))
        card_names = list(CreditCard.objects.filter(tenant_id=self.tenant_id).order_by('name').values_list('name', flat=True))
        headers = ['descricao', 'tipo', 'valor', 'data', 'categoria', 'metodo_pagamento', 'quantidade', 'parcelado', 'parcelas', 'cartao']
        wb = Workbook()
        ws = wb.active
        ws.title = 'Gastos'
        bold = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold
        sample_category = categories[0] if categories else 'Alimenta??o'
        ws.append(['Mercado do m?s', 'despesa', 150.00, '2026-08-01', sample_category, 'dinheiro', 1, 'nao', 1, ''])
        widths = [28, 12, 12, 14, 22, 18, 12, 12, 10, 22]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = 'A2'
        cat_ws = wb.create_sheet('Categorias')
        cat_ws.cell(row=1, column=1, value='Categorias dispon?veis').font = bold
        for i, name in enumerate(categories, start=2):
            cat_ws.cell(row=i, column=1, value=name)
        tipo_dv = DataValidation(type='list', formula1='"despesa,receita"', allow_blank=False)
        ws.add_data_validation(tipo_dv)
        tipo_dv.add('B2:B1000')
        if categories:
            last = len(categories) + 1
            cat_dv = DataValidation(type='list', formula1=f'=Categorias!$A$2:$A${last}', allow_blank=True)
            ws.add_data_validation(cat_dv)
            cat_dv.add('E2:E1000')
        pm_dv = DataValidation(type='list', formula1='"dinheiro,cartao"', allow_blank=True)
        ws.add_data_validation(pm_dv)
        pm_dv.add('F2:F1000')
        parc_dv = DataValidation(type='list', formula1='"sim,nao"', allow_blank=True)
        ws.add_data_validation(parc_dv)
        parc_dv.add('H2:H1000')
        if card_names:
            card_ws = wb.create_sheet('Cartoes')
            card_ws.cell(row=1, column=1, value='Cart?es dispon?veis').font = bold
            for i, name in enumerate(card_names, start=2):
                card_ws.cell(row=i, column=1, value=name)
            card_dv = DataValidation(type='list', formula1=f'=Cartoes!$A$2:$A${len(card_names) + 1}', allow_blank=True)
            ws.add_data_validation(card_dv)
            card_dv.add('J2:J1000')
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="template_gastos.xlsx"'
        return resp

    @staticmethod
    def _get_cell(row, header_map, name):
        idx = header_map.get(name)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    @staticmethod
    def _parse_date(value):
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date_cls):
            return value
        s = str(value).strip()
        for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def import_excel(self, upload) -> Response:
        """Importa gastos de uma planilha .xlsx com semantica tudo-ou-nada."""
        if upload is None:
            return Response({'success': False, 'message': 'Arquivo n?o enviado'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(upload, data_only=True)
        except Exception:
            return Response({'success': False, 'message': 'Arquivo inv?lido ou corrompido'}, status=status.HTTP_400_BAD_REQUEST)
        ws = wb['Gastos'] if 'Gastos' in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'success': False, 'message': 'Planilha vazia'}, status=status.HTTP_400_BAD_REQUEST)
        header_map = {str(val).strip().lower(): idx for idx, val in enumerate(rows[0]) if val is not None}
        missing = {'descricao', 'valor', 'data', 'categoria'} - set(header_map.keys())
        if missing:
            return Response({'success': False, 'message': f'Colunas obrigat?rias ausentes: {sorted(missing)}'}, status=status.HTTP_400_BAD_REQUEST)
        data_rows = rows[1:]
        if len(data_rows) > self.MAX_ROWS:
            return Response({'success': False, 'message': f'N?mero de linhas excede o m?ximo de {self.MAX_ROWS}'}, status=status.HTTP_400_BAD_REQUEST)
        category_map = {c.name.strip(): c.id for c in ExpenseCategory.objects.filter(tenant_id__in=['system', self.tenant_id])}
        from cards.models import CreditCard
        card_map = {c.name.strip().lower(): c.id for c in CreditCard.objects.filter(tenant_id=self.tenant_id)}
        errors, parsed_items = [], []
        for offset, row in enumerate(data_rows):
            excel_row = offset + 2
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            row_errors = []
            get_cell = lambda name: self._get_cell(row, header_map, name)
            descricao = get_cell('descricao')
            descricao = str(descricao).strip() if descricao is not None else ''
            if not descricao:
                row_errors.append('descricao vazia')
            elif len(descricao) > 255:
                row_errors.append('descricao excede 255 caracteres')
            valor_raw = get_cell('valor')
            amount = None
            try:
                amount = Decimal(str(valor_raw).strip().replace(',', '.'))
            except (InvalidOperation, AttributeError, TypeError):
                row_errors.append('valor inv?lido')
            tipo_raw = get_cell('tipo')
            tipo = str(tipo_raw).strip().lower() if tipo_raw not in (None, '') else 'despesa'
            if tipo not in ('despesa', 'receita'):
                row_errors.append('tipo deve ser despesa/receita')
            if amount is not None and tipo in ('despesa', 'receita'):
                amount = abs(amount) if tipo == 'receita' else -abs(amount)
            data_val = self._parse_date(get_cell('data'))
            if data_val is None:
                row_errors.append('data inv?lida')
            cat_name = get_cell('categoria')
            cat_name = str(cat_name).strip() if cat_name is not None else ''
            category_id = category_map.get(cat_name)
            if category_id is None:
                row_errors.append(f'categoria "{cat_name}" n?o encontrada')
            pm_raw = get_cell('metodo_pagamento')
            payment_method = str(pm_raw).strip().lower() if pm_raw not in (None, '') else 'dinheiro'
            if payment_method not in ('dinheiro', 'cartao'):
                row_errors.append('metodo_pagamento inv?lido')
            card_raw = get_cell('cartao')
            card_name = str(card_raw).strip() if card_raw not in (None, '') else ''
            credit_card_id = None
            if payment_method == 'cartao':
                if not card_name:
                    row_errors.append('cartao ? obrigat?rio quando metodo_pagamento = "cartao"')
                else:
                    credit_card_id = card_map.get(card_name.lower())
                    if credit_card_id is None:
                        row_errors.append(f'cartao "{card_name}" n?o encontrado')
            elif card_name:
                row_errors.append('cartao s? pode ser informado quando metodo_pagamento = "cartao"')
            qty_raw = get_cell('quantidade')
            quantity = 1
            if qty_raw not in (None, ''):
                try:
                    quantity = int(qty_raw)
                except (ValueError, TypeError):
                    row_errors.append('quantidade inv?lida')
            if quantity < 1:
                row_errors.append('quantidade deve ser >= 1')
            parc_raw = get_cell('parcelado')
            parcelado = str(parc_raw).strip().lower() if parc_raw not in (None, '') else 'nao'
            if parcelado not in ('sim', 'nao'):
                row_errors.append('parcelado deve ser sim/nao')
            parcelas_raw = get_cell('parcelas')
            parcelas = 1
            if parcelas_raw not in (None, ''):
                try:
                    parcelas = int(parcelas_raw)
                except (ValueError, TypeError):
                    row_errors.append('parcelas inv?lida')
            if parcelas < 1:
                row_errors.append('parcelas deve ser >= 1')
            if row_errors:
                errors.append({'row': excel_row, 'error': '; '.join(row_errors)})
                continue
            parsed_items.append({'category_id': category_id, 'description': descricao, 'amount': amount, 'date': data_val, 'quantity': quantity, 'payment_method': payment_method, 'credit_card_id': credit_card_id, 'is_installment': parcelado == 'sim', 'installments': parcelas})
        if errors:
            return Response({'success': False, 'errors': errors, 'message': f'{len(errors)} linha(s) com erro; nada foi importado'}, status=status.HTTP_400_BAD_REQUEST)
        if not parsed_items:
            return Response({'success': False, 'message': 'Nenhuma linha de dados encontrada'}, status=status.HTTP_400_BAD_REQUEST)
        created = []
        try:
            with transaction.atomic():
                for item in parsed_items:
                    created.extend(self._create_from_item(item))
        except Exception as e:
            return Response({'success': False, 'message': f'Erro ao importar gastos: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)
        return Response({'success': True, 'created': len(created), 'message': f'{len(created)} gasto(s) importado(s) com sucesso'}, status=status.HTTP_201_CREATED)

    def delete_installments(self, description_prefix, total_installments) -> Response:
        """Remove todas as parcelas de uma despesa parcelada de uma so vez."""
        pattern = rf'^{re.escape(description_prefix)} - Parcela \d+/{total_installments}$'
        qs = models.Expense.objects.filter(tenant_id=self.tenant_id, description__iregex=pattern)
        with transaction.atomic():
            deleted_count, _ = qs.delete()
        if deleted_count == 0:
            return Response({'error': 'Nenhuma parcela encontrada para os crit?rios informados.'}, status=status.HTTP_404_NOT_FOUND)
        return Response({'deleted': deleted_count, 'description_prefix': description_prefix, 'total_installments': total_installments}, status=status.HTTP_200_OK)
