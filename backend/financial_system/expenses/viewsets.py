import re
from datetime import date

from django.db import transaction
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response

from expenses import models, serializer
from expenses.behaviors import CreateExpenseBehavior, RecurringExpenseBehavior
from expenses.analytics_behaviors import ExpenseAnalyticsBehavior
from catalog.models import ExpenseCategory


class ExpensePagination(PageNumberPagination):
    """
    Paginação padrão para o endpoint de expenses.

    Query params disponíveis:
      - page       : número da página (começa em 1)
      - page_size  : tamanho da página (padrão 20, máximo 100)
    """

    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100
    page_query_param = 'page'


class ExpenseViewSet(viewsets.ModelViewSet):
    serializer_class = serializer.ExpenseSerializer
    queryset = models.Expense.objects.all()
    pagination_class = ExpensePagination

    # ------------------------------------------------------------------
    # Query params suportados pelo endpoint de listagem (GET /expenses/):
    #   month          (int 1-12)  – filtra por mês da data
    #   year           (int)       – filtra por ano da data
    #   category_id    (int)       – filtra por categoria
    #   payment_method (str)       – "dinheiro" | "cartao"
    #   search         (str)       – busca case-insensitive em description
    #
    # month e year são independentes: podem ser usados individualmente ou
    # em conjunto. Quando nenhum dos dois for informado, nenhum filtro de
    # data é aplicado (suporte a buscas históricas).
    # ------------------------------------------------------------------

    def get_queryset(self):
        qs = (
            models.Expense.objects
            .filter(tenant_id=self.request.user.tenant_id)
            # Evita N+1 ao serializar category e credit_card em cada registro
            .select_related('category', 'credit_card')
            .order_by('-date', '-id')
        )

        params = self.request.query_params

        # --- month ---------------------------------------------------
        raw_month = params.get('month')
        if raw_month is not None:
            try:
                month = int(raw_month)
                if 1 <= month <= 12:
                    qs = qs.filter(date__month=month)
            except (ValueError, TypeError):
                pass  # parâmetro inválido ignorado silenciosamente

        # --- year ----------------------------------------------------
        raw_year = params.get('year')
        if raw_year is not None:
            try:
                year = int(raw_year)
                if year > 0:
                    qs = qs.filter(date__year=year)
            except (ValueError, TypeError):
                pass

        # --- category_id ---------------------------------------------
        raw_category = params.get('category_id')
        if raw_category is not None:
            try:
                category_id = int(raw_category)
                qs = qs.filter(category_id=category_id)
            except (ValueError, TypeError):
                pass

        # --- payment_method ------------------------------------------
        payment_method = params.get('payment_method')
        if payment_method is not None:
            valid_choices = {choice[0] for choice in models.Expense.PAYMENT_METHOD_CHOICES}
            if payment_method in valid_choices:
                qs = qs.filter(payment_method=payment_method)

        # --- credit_card_id ------------------------------------------
        raw_credit_card = params.get('credit_card_id')
        if raw_credit_card is not None:
            try:
                qs = qs.filter(credit_card_id=int(raw_credit_card))
            except (ValueError, TypeError):
                pass

        # --- start_date / end_date -----------------------------------
        raw_start_date = params.get('start_date')
        if raw_start_date:
            try:
                qs = qs.filter(date__gte=date.fromisoformat(raw_start_date))
            except (ValueError, TypeError):
                pass

        raw_end_date = params.get('end_date')
        if raw_end_date:
            try:
                qs = qs.filter(date__lte=date.fromisoformat(raw_end_date))
            except (ValueError, TypeError):
                pass

        # --- search (description icontains) --------------------------
        # [SEC-A03] Limita o tamanho da query para prevenir DoS via queries longas
        # (icontains é parameterizado pelo ORM — seguro contra SQL injection)
        search = params.get('search')
        if search:
            search = search.strip()[:200]
            if search:
                qs = qs.filter(description__icontains=search)

        return qs

    def perform_update(self, serializer):
        """[SEC-A01] Defense-in-depth: garante que tenant_id não muda em updates."""
        serializer.save(tenant_id=self.request.user.tenant_id)

    # ------------------------------------------------------------------
    # Custom actions — CRUD helpers
    # ------------------------------------------------------------------

    def perform_create(self, serializer):
        """Injeta tenant_id do usuário autenticado ao criar via POST padrão."""
        serializer.save(tenant_id=self.request.user.tenant_id)

    def perform_destroy(self, instance):
        if instance.tenant_id != self.request.user.tenant_id:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Você não tem permissão para excluir este recurso.")
        instance.delete()

    @action(detail=False, methods=['post'], url_path='create-expense')
    def create_expense(self, request):
        s = serializer.CreateExpenseInputSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        payload = dict(s.validated_data)
        payload['tenant_id'] = request.user.tenant_id
        return CreateExpenseBehavior(data=payload).run()

    def _create_from_item(self, item: dict, tenant_id):
        """Cria despesa(s) a partir de um item validado, respeitando
        quantidade/parcelamento. Retorna a lista de objetos criados."""
        payload = dict(item)
        payload['tenant_id'] = tenant_id
        behavior = CreateExpenseBehavior(data=payload)
        if behavior.is_installment and behavior.installments > 1:
            return behavior._create_installments()
        if behavior.quantity and behavior.quantity > 1:
            return behavior._create_multiple()
        return [behavior._create_single()]

    @action(detail=False, methods=['post'], url_path='bulk-create')
    def bulk_create(self, request):
        """
        POST /api/expenses/expenses/bulk-create/

        Cria múltiplos gastos de forma atômica (tudo ou nada). Cada item segue
        a mesma validação de create-expense (CreateExpenseInputSerializer).
        """
        s = serializer.BulkCreateExpenseInputSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        items = s.validated_data['items']
        tenant = request.user.tenant_id

        # [SEC] Valida propriedade das categorias ANTES de qualquer DML —
        # impede referenciar categoria de outro tenant.
        requested_ids = {item['category_id'] for item in items}
        allowed_ids = set(
            ExpenseCategory.objects.filter(
                tenant_id__in=['system', tenant],
                id__in=requested_ids,
            ).values_list('id', flat=True)
        )
        invalid_ids = sorted(requested_ids - allowed_ids)
        if invalid_ids:
            return Response(
                {
                    'success': False,
                    'message': f'Categoria(s) inválida(s): {invalid_ids}',
                    'invalid_category_ids': invalid_ids,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = []
        try:
            with transaction.atomic():
                for item in items:
                    created.extend(self._create_from_item(item, tenant))
        except Exception as e:
            return Response(
                {'success': False, 'message': f'Erro ao criar gastos em lote: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {
                'success': True,
                'created': len(created),
                'message': f'{len(created)} gasto(s) criado(s) com sucesso',
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['get'], url_path='import-template')
    def import_template(self, request):
        """
        GET /api/expenses/expenses/import-template/

        Retorna um arquivo .xlsx modelo para importação de gastos, com
        dropdowns de categoria (válidas para o tenant), método de pagamento
        e parcelado.
        """
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation

        tenant = request.user.tenant_id
        categories = list(
            ExpenseCategory.objects
            .filter(tenant_id__in=['system', tenant])
            .order_by('name')
            .values_list('name', flat=True)
        )

        from cards.models import CreditCard
        card_names = list(
            CreditCard.objects
            .filter(tenant_id=tenant)
            .order_by('name')
            .values_list('name', flat=True)
        )

        headers = [
            'descricao', 'tipo', 'valor', 'data', 'categoria',
            'metodo_pagamento', 'quantidade', 'parcelado', 'parcelas',
            'cartao',
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = 'Gastos'

        bold = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = bold

        # Linha de exemplo (valor POSITIVO — o sinal vem de `tipo`)
        sample_category = categories[0] if categories else 'Alimentação'
        ws.append([
            'Mercado do mês', 'despesa', 150.00, '2026-08-01', sample_category,
            'dinheiro', 1, 'nao', 1, '',
        ])

        # Larguras razoáveis (10 colunas)
        widths = [28, 12, 12, 14, 22, 18, 12, 12, 10, 22]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.freeze_panes = 'A2'

        # Sheet de categorias (fonte para o dropdown)
        cat_ws = wb.create_sheet('Categorias')
        cat_ws.cell(row=1, column=1, value='Categorias disponíveis').font = bold
        for i, name in enumerate(categories, start=2):
            cat_ws.cell(row=i, column=1, value=name)

        # Colunas: A descricao, B tipo, C valor, D data, E categoria,
        #          F metodo_pagamento, G quantidade, H parcelado, I parcelas,
        #          J cartao (obrigatório quando metodo_pagamento = "cartao").

        # Dropdown tipo (coluna B)
        tipo_dv = DataValidation(
            type='list', formula1='"despesa,receita"', allow_blank=False,
        )
        ws.add_data_validation(tipo_dv)
        tipo_dv.add('B2:B1000')

        # Dropdown de categoria (coluna E). Se não houver categorias, pula.
        if categories:
            last = len(categories) + 1
            cat_dv = DataValidation(
                type='list',
                formula1=f'=Categorias!$A$2:$A${last}',
                allow_blank=True,
            )
            ws.add_data_validation(cat_dv)
            cat_dv.add('E2:E1000')

        # Dropdown método de pagamento (coluna F)
        pm_dv = DataValidation(
            type='list', formula1='"dinheiro,cartao"', allow_blank=True,
        )
        ws.add_data_validation(pm_dv)
        pm_dv.add('F2:F1000')

        # Dropdown parcelado (coluna H)
        parc_dv = DataValidation(
            type='list', formula1='"sim,nao"', allow_blank=True,
        )
        ws.add_data_validation(parc_dv)
        parc_dv.add('H2:H1000')

        # Sheet + dropdown de cartões (coluna J).
        if card_names:
            card_ws = wb.create_sheet('Cartoes')
            card_ws.cell(row=1, column=1, value='Cartões disponíveis').font = bold
            for i, name in enumerate(card_names, start=2):
                card_ws.cell(row=i, column=1, value=name)
            card_dv = DataValidation(
                type='list',
                formula1=f'=Cartoes!$A$2:$A${len(card_names) + 1}',
                allow_blank=True,
            )
            ws.add_data_validation(card_dv)
            card_dv.add('J2:J1000')

        from django.http import HttpResponse
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(
            bio.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        resp['Content-Disposition'] = 'attachment; filename="template_gastos.xlsx"'
        return resp

    @action(detail=False, methods=['post'], url_path='import-excel')
    def import_excel(self, request):
        """
        POST /api/expenses/expenses/import-excel/  (multipart/form-data)

        Importa gastos a partir de um arquivo .xlsx (campo 'file').
        Semântica tudo-ou-nada: se qualquer linha tiver erro, nada é criado.
        """
        from datetime import date as date_cls, datetime
        from decimal import Decimal, InvalidOperation

        MAX_ROWS = 1000

        upload = request.FILES.get('file')
        if upload is None:
            return Response(
                {'success': False, 'message': 'Arquivo não enviado'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from openpyxl import load_workbook
            wb = load_workbook(upload, data_only=True)
        except Exception:
            return Response(
                {'success': False, 'message': 'Arquivo inválido ou corrompido'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        ws = wb['Gastos'] if 'Gastos' in wb.sheetnames else wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response(
                {'success': False, 'message': 'Planilha vazia'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Mapa header -> índice (resiliente à ordem das colunas)
        header_row = rows[0]
        header_map = {}
        for idx, val in enumerate(header_row):
            if val is not None:
                header_map[str(val).strip().lower()] = idx

        required_headers = {'descricao', 'valor', 'data', 'categoria'}
        missing = required_headers - set(header_map.keys())
        if missing:
            return Response(
                {'success': False, 'message': f'Colunas obrigatórias ausentes: {sorted(missing)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        data_rows = rows[1:]
        if len(data_rows) > MAX_ROWS:
            return Response(
                {'success': False, 'message': f'Número de linhas excede o máximo de {MAX_ROWS}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        tenant = request.user.tenant_id
        category_map = {
            c.name.strip(): c.id
            for c in ExpenseCategory.objects.filter(tenant_id__in=['system', tenant])
        }

        from cards.models import CreditCard
        card_map = {
            c.name.strip().lower(): c.id
            for c in CreditCard.objects.filter(tenant_id=tenant)
        }

        def get_cell(row, name):
            idx = header_map.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        def parse_date(value):
            if isinstance(value, datetime):
                return value.date()
            if isinstance(value, date_cls):
                return value
            s = str(value).strip()
            for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
            return None

        errors = []
        parsed_items = []

        for offset, row in enumerate(data_rows):
            excel_row = offset + 2  # header é linha 1
            # Ignora linhas totalmente vazias
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue

            row_errors = []

            descricao = get_cell(row, 'descricao')
            descricao = str(descricao).strip() if descricao is not None else ''
            if not descricao:
                row_errors.append('descricao vazia')
            elif len(descricao) > 255:
                row_errors.append('descricao excede 255 caracteres')

            valor_raw = get_cell(row, 'valor')
            amount = None
            try:
                amount = Decimal(str(valor_raw).strip().replace(',', '.'))
            except (InvalidOperation, AttributeError, TypeError):
                row_errors.append('valor inválido')

            tipo_raw = get_cell(row, 'tipo')
            tipo = str(tipo_raw).strip().lower() if tipo_raw not in (None, '') else 'despesa'
            if tipo not in ('despesa', 'receita'):
                row_errors.append('tipo deve ser despesa/receita')

            # Aplica o sinal a partir do tipo (magnitude absoluta), garantindo
            # despesa negativa e receita positiva independente do que o usuário digitou.
            if amount is not None and tipo in ('despesa', 'receita'):
                amount = abs(amount) if tipo == 'receita' else -abs(amount)

            data_val = parse_date(get_cell(row, 'data'))
            if data_val is None:
                row_errors.append('data inválida')

            cat_name = get_cell(row, 'categoria')
            cat_name = str(cat_name).strip() if cat_name is not None else ''
            category_id = category_map.get(cat_name)
            if category_id is None:
                row_errors.append(f'categoria "{cat_name}" não encontrada')

            pm_raw = get_cell(row, 'metodo_pagamento')
            payment_method = str(pm_raw).strip().lower() if pm_raw not in (None, '') else 'dinheiro'
            if payment_method not in ('dinheiro', 'cartao'):
                row_errors.append('metodo_pagamento inválido')

            # cartao sem credit_card vinculado gera lançamento invisível na fatura.
            card_raw = get_cell(row, 'cartao')
            card_name = str(card_raw).strip() if card_raw not in (None, '') else ''
            credit_card_id = None
            if payment_method == 'cartao':
                if not card_name:
                    row_errors.append(
                        'cartao é obrigatório quando metodo_pagamento = "cartao"'
                    )
                else:
                    credit_card_id = card_map.get(card_name.lower())
                    if credit_card_id is None:
                        row_errors.append(f'cartao "{card_name}" não encontrado')
            elif card_name:
                row_errors.append(
                    'cartao só pode ser informado quando metodo_pagamento = "cartao"'
                )

            qty_raw = get_cell(row, 'quantidade')
            quantity = 1
            if qty_raw not in (None, ''):
                try:
                    quantity = int(qty_raw)
                except (ValueError, TypeError):
                    row_errors.append('quantidade inválida')
            if quantity < 1:
                row_errors.append('quantidade deve ser >= 1')

            parc_raw = get_cell(row, 'parcelado')
            parcelado = str(parc_raw).strip().lower() if parc_raw not in (None, '') else 'nao'
            if parcelado not in ('sim', 'nao'):
                row_errors.append('parcelado deve ser sim/nao')

            parcelas_raw = get_cell(row, 'parcelas')
            parcelas = 1
            if parcelas_raw not in (None, ''):
                try:
                    parcelas = int(parcelas_raw)
                except (ValueError, TypeError):
                    row_errors.append('parcelas inválida')
            if parcelas < 1:
                row_errors.append('parcelas deve ser >= 1')

            if row_errors:
                errors.append({'row': excel_row, 'error': '; '.join(row_errors)})
                continue

            parsed_items.append({
                'category_id': category_id,
                'description': descricao,
                'amount': amount,
                'date': data_val,
                'quantity': quantity,
                'payment_method': payment_method,
                'credit_card_id': credit_card_id,
                'is_installment': parcelado == 'sim',
                'installments': parcelas,
            })

        if errors:
            return Response(
                {
                    'success': False,
                    'errors': errors,
                    'message': f'{len(errors)} linha(s) com erro; nada foi importado',
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not parsed_items:
            return Response(
                {'success': False, 'message': 'Nenhuma linha de dados encontrada'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        created = []
        try:
            with transaction.atomic():
                for item in parsed_items:
                    created.extend(self._create_from_item(item, tenant))
        except Exception as e:
            return Response(
                {'success': False, 'message': f'Erro ao importar gastos: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        return Response(
            {
                'success': True,
                'created': len(created),
                'message': f'{len(created)} gasto(s) importado(s) com sucesso',
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['post'], url_path='delete-installments')
    def delete_installments(self, request):
        """
        POST /api/expenses/expenses/delete-installments/

        Remove todas as parcelas de uma despesa parcelada de uma só vez.

        O padrão de descrição esperado é o gerado por CreateExpenseBehavior:
          "{description_prefix} - Parcela {X}/{total_installments}"

        Request body:
          - description_prefix   (str, obrigatório, max 255 chars)
          - total_installments   (int, obrigatório, mínimo 2)

        Responses:
          200  {"deleted": N, "description_prefix": "...", "total_installments": N}
          404  {"error": "Nenhuma parcela encontrada para os critérios informados."}
          400  DRF validation errors
        """
        s = serializer.DeleteInstallmentsInputSerializer(data=request.data)
        s.is_valid(raise_exception=True)

        description_prefix = s.validated_data['description_prefix']
        total_installments = s.validated_data['total_installments']

        # Build a regex that matches exactly the installment format produced by
        # CreateExpenseBehavior: "{BaseName} - Parcela {X}/{total}"
        # re.escape ensures user-supplied characters (e.g. dots, parens) are safe.
        pattern = rf'^{re.escape(description_prefix)} - Parcela \d+/{total_installments}$'

        # [SEC-A01] Always scope to the authenticated tenant before any DML.
        qs = models.Expense.objects.filter(
            tenant_id=request.user.tenant_id,
            description__iregex=pattern,
        )

        with transaction.atomic():
            deleted_count, _ = qs.delete()

        if deleted_count == 0:
            return Response(
                {'error': 'Nenhuma parcela encontrada para os critérios informados.'},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(
            {
                'deleted': deleted_count,
                'description_prefix': description_prefix,
                'total_installments': total_installments,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=['get'], url_path='per-credit-card/(?P<card_id>[0-9]+)')
    def expenses_per_credit_card(self, request, card_id=None):
        """
        Retorna todas as despesas de um cartão de crédito sem paginação
        (lista simples — necessária para cálculos de fatura completa).
        """
        qs = (
            models.Expense.objects
            .filter(
                tenant_id=request.user.tenant_id,
                credit_card_id=card_id,
            )
            # Evita N+1 ao serializar category e credit_card
            .select_related('category', 'credit_card')
            .order_by('-date', '-id')
        )
        s = serializer.ExpenseSerializer(qs, many=True)
        return Response(s.data, status=status.HTTP_200_OK)

    # ------------------------------------------------------------------
    # Analytics actions
    # ------------------------------------------------------------------

    @action(detail=False, methods=['get'], url_path='analytics/monthly')
    def analytics_monthly(self, request):
        """GET analytics/monthly."""
        data = ExpenseAnalyticsBehavior(request.user.tenant_id).analytics_monthly(
            request.query_params
        )
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='analytics/by-category')
    def analytics_by_category(self, request):
        """GET analytics/by-category."""
        data = ExpenseAnalyticsBehavior(request.user.tenant_id).analytics_by_category(
            request.query_params
        )
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='analytics/by-card')
    def analytics_by_card(self, request):
        """GET analytics/by-card."""
        data = ExpenseAnalyticsBehavior(request.user.tenant_id).analytics_by_card(
            request.query_params
        )
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='analytics/daily')
    def analytics_daily(self, request):
        """GET analytics/daily."""
        data = ExpenseAnalyticsBehavior(request.user.tenant_id).analytics_daily(
            request.query_params
        )
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='consolidated-summary')
    def consolidated_summary(self, request):
        """GET consolidated-summary."""
        data = ExpenseAnalyticsBehavior(request.user.tenant_id).consolidated_summary(
            request.query_params
        )
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], url_path='home-charts')
    def home_charts(self, request):
        """GET home-charts."""
        data = ExpenseAnalyticsBehavior(request.user.tenant_id).home_charts(
            request.query_params
        )
        return Response(data, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get', 'post'], url_path='recurring-templates')
    def recurring_templates(self, request):
        """GET: listar templates, POST: criar template"""
        behavior = RecurringExpenseBehavior(request.user.tenant_id)
        if request.method == 'GET':
            return behavior.list()
        s = serializer.CreateRecurringExpenseInputSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        return behavior.create({**dict(s.validated_data), 'tenant_id': request.user.tenant_id})

    @action(detail=False, methods=['delete', 'patch', 'put'],
            url_path=r'recurring-templates/(?P<tpl_id>[0-9]+)')
    def recurring_template_detail(self, request, tpl_id=None):
        """DELETE: excluir, PATCH: toggle is_active, PUT: editar"""
        behavior = RecurringExpenseBehavior(request.user.tenant_id)
        if request.method == 'DELETE':
            return behavior.delete(int(tpl_id))
        if request.method == 'PUT':
            s = serializer.CreateRecurringExpenseInputSerializer(data=request.data)
            s.is_valid(raise_exception=True)
            return behavior.update(int(tpl_id), dict(s.validated_data))
        return behavior.toggle_active(int(tpl_id))

    @action(detail=False, methods=['post'], url_path='recurring-templates/generate-month')
    def generate_month_recurring(self, request):
        """POST {month, year} → materializa todos os templates ativos"""
        s = serializer.GenerateMonthInputSerializer(data=request.data)
        s.is_valid(raise_exception=True)
        return RecurringExpenseBehavior(request.user.tenant_id).generate_month(
            s.validated_data['month'], s.validated_data['year']
        )

